#!/usr/bin/env python3
"""Baixa PDFs open access e atualiza o catálogo de artigos."""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import unicodedata
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import quote, urlencode, urljoin, urlparse
from urllib.request import Request, urlopen

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PLANILHA = ROOT / "bibliography/papers/catalogo_artigos.xlsx"
PASTA_PDFS = ROOT / "bibliography/papers"
BIB_PATHS = (
    ROOT / "bibliography/reference.bib",
    ROOT / "bibliography/new_papers_2022_2025.bib",
)
NOTAS_DIR = ROOT / "bibliography/papers/notes"

ABA_PADRAO = "papers_useful"
COLUNA_TITULO = "Título do artigo"
COLUNA_ANO = "Ano de publicação"
COLUNA_AUTORES = "Autores"
COLUNA_BAIXADO = "Artigo baixado no computador?"
NOVAS_COLUNAS = (
    "DOI",
    "Status do download OA",
    "URL do PDF OA",
    "Arquivo local OA",
)
STATUS_BAIXADO = "baixado"
STATUS_EXISTENTE = "arquivo já existe"
STATUS_MANUAL = "manual"
STATUS_SEM_DOI = "sem DOI"
STATUS_SEM_PDF_OA = "sem PDF open access"
STATUS_ERRO = "erro"
USER_AGENT = "lightning-bibliografia-oa/1.0"
BROWSER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7",
    "Cache-Control": "no-cache",
}
STOPWORDS = {
    "a",
    "an",
    "and",
    "as",
    "at",
    "by",
    "da",
    "das",
    "de",
    "do",
    "dos",
    "for",
    "from",
    "in",
    "into",
    "its",
    "of",
    "on",
    "or",
    "over",
    "the",
    "their",
    "to",
    "under",
    "using",
    "via",
    "with",
}


@dataclass
class LinhaArtigo:
    aba: str
    indice_linha: int
    titulo: str
    autores: str
    ano: int | None
    doi: str


@dataclass
class PdfLocal:
    caminho: Path
    stem_norm: str
    titulo_metadata_norm: str
    primeira_pagina_norm: str


def normalizar(texto: str | None) -> str:
    if texto is None:
        return ""
    texto = (
        unicodedata.normalize("NFKD", str(texto))
        .encode("ascii", "ignore")
        .decode("ascii")
        .lower()
    )
    texto = re.sub(r"[^a-z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def limpar_doi(texto: str | None) -> str:
    if not texto:
        return ""
    doi = str(texto).strip()
    doi = doi.replace("https://doi.org/", "").replace("http://doi.org/", "")
    doi = doi.replace("doi:", "").strip()
    return doi


def autor_principal(texto: str | None) -> str:
    if not texto:
        return "desconhecido"
    bruto = str(texto).strip()
    primeiro = re.split(r";|, and | and | et al\.?", bruto, maxsplit=1)[0].strip()
    tokens = re.findall(r"[A-Za-zÀ-ÿ]+", primeiro)
    if not tokens:
        return "desconhecido"
    return normalizar(tokens[-1]) or "desconhecido"


def token_recall(referencia: str, alvo: str) -> float:
    ref_tokens = [t for t in referencia.split() if t not in STOPWORDS and not t.isdigit()]
    if not ref_tokens:
        return 0.0
    alvo_tokens = {t for t in alvo.split() if t not in STOPWORDS and not t.isdigit()}
    return sum(1 for token in ref_tokens if token in alvo_tokens) / len(ref_tokens)


def titulo_metadata_pdf(pdf: Path) -> str:
    try:
        saida = subprocess.run(
            ["pdfinfo", str(pdf)],
            capture_output=True,
            text=True,
            check=False,
        ).stdout
    except Exception:
        return ""
    match = re.search(r"^Title:\s*(.*)$", saida, flags=re.MULTILINE)
    return normalizar(match.group(1).strip()) if match else ""


def primeira_pagina_pdf(pdf: Path) -> str:
    try:
        saida = subprocess.run(
            ["pdftotext", "-f", "1", "-l", "1", "-nopgbrk", str(pdf), "-"],
            capture_output=True,
            text=True,
            check=False,
        ).stdout
    except Exception:
        return ""
    return normalizar(saida)


def inventariar_pdfs_locais() -> list[PdfLocal]:
    inventario: list[PdfLocal] = []
    for pdf in PASTA_PDFS.rglob("*.pdf"):
        inventario.append(
            PdfLocal(
                caminho=pdf,
                stem_norm=normalizar(pdf.stem),
                titulo_metadata_norm=titulo_metadata_pdf(pdf),
                primeira_pagina_norm=primeira_pagina_pdf(pdf),
            )
        )
    return inventario


def localizar_pdf_local_existente(titulo: str, inventario: list[PdfLocal]) -> Path | None:
    titulo_norm = normalizar(titulo)
    if not titulo_norm:
        return None

    melhor_score = 0.0
    melhor_caminho: Path | None = None

    for pdf in inventario:
        if titulo_norm and titulo_norm in pdf.primeira_pagina_norm and pdf.primeira_pagina_norm:
            return pdf.caminho

        for alvo in (pdf.stem_norm, pdf.titulo_metadata_norm):
            if not alvo:
                continue
            score = max(
                token_recall(titulo_norm, alvo),
                SequenceMatcher(None, titulo_norm, alvo).ratio(),
            )
            if score > melhor_score:
                melhor_score = score
                melhor_caminho = pdf.caminho

    return melhor_caminho if melhor_score >= 0.72 else None


def request_json(url: str, params: dict[str, Any] | None = None) -> dict[str, Any]:
    if params:
        url = f"{url}?{urlencode(params)}"
    request = Request(
        url,
        headers={
            **BROWSER_HEADERS,
            "User-Agent": USER_AGENT,
            "Accept": "application/json",
        },
    )
    with urlopen(request, timeout=30) as response:
        return json.load(response)


def split_bib_entries(texto: str) -> list[str]:
    entradas: list[str] = []
    atual: list[str] = []
    for linha in texto.splitlines():
        if linha.lstrip().startswith("@") and atual:
            entradas.append("\n".join(atual))
            atual = [linha]
        else:
            atual.append(linha)
    if atual:
        entradas.append("\n".join(atual))
    return entradas


def extrair_campo_bib(entry: str, campo: str) -> str:
    padroes = (
        rf"{campo}\s*=\s*\{{(.*?)\}}",
        rf'{campo}\s*=\s*"(.*?)"',
    )
    for padrao in padroes:
        match = re.search(padrao, entry, flags=re.IGNORECASE | re.DOTALL)
        if match:
            valor = re.sub(r"\s+", " ", match.group(1)).strip()
            return valor
    return ""


def carregar_dois_locais() -> dict[str, str]:
    indice: dict[str, str] = {}

    for bib_path in BIB_PATHS:
        if not bib_path.exists():
            continue
        for entry in split_bib_entries(bib_path.read_text(encoding="utf-8", errors="ignore")):
            titulo = extrair_campo_bib(entry, "title")
            doi = limpar_doi(extrair_campo_bib(entry, "doi"))
            if titulo and doi:
                indice[normalizar(titulo)] = doi

    if NOTAS_DIR.exists():
        for nota in NOTAS_DIR.glob("*.md"):
            conteudo = nota.read_text(encoding="utf-8", errors="ignore")
            doi_match = re.search(r"10\.\d{4,9}/[^\s\}\]]+", conteudo, flags=re.IGNORECASE)
            if not doi_match:
                continue
            doi = limpar_doi(doi_match.group(0))
            titulo = ""
            full_title = re.search(r"\*\*Full title\*\*\s*\n([^\n]+)", conteudo, flags=re.IGNORECASE)
            if full_title:
                titulo = full_title.group(1).strip()
            if not titulo:
                titulo = extrair_campo_bib(conteudo, "title")
            if titulo:
                indice.setdefault(normalizar(titulo), doi)

    return indice


def buscar_doi_crossref(titulo: str, ano: int | None, mailto: str | None) -> str:
    params: dict[str, Any] = {
        "query.title": titulo,
        "rows": 5,
        "select": "DOI,title,published,published-print,published-online",
    }
    if mailto:
        params["mailto"] = mailto
    try:
        payload = request_json("https://api.crossref.org/works", params=params)
    except (HTTPError, URLError, TimeoutError, json.JSONDecodeError):
        return ""

    candidatos = payload.get("message", {}).get("items", [])
    titulo_norm = normalizar(titulo)
    melhor_doi = ""
    melhor_score = 0.0

    for item in candidatos:
        titulos = item.get("title") or []
        if not titulos:
            continue
        candidato_titulo = titulos[0]
        candidato_norm = normalizar(candidato_titulo)
        similaridade = SequenceMatcher(None, titulo_norm, candidato_norm).ratio()
        cobertura = token_recall(titulo_norm, candidato_norm)
        score = max(similaridade, cobertura)

        ano_candidato = None
        for campo in ("published-print", "published-online", "published"):
            partes = item.get(campo, {}).get("date-parts") or []
            if partes and partes[0]:
                ano_candidato = partes[0][0]
                break

        if ano and ano_candidato == ano:
            score += 0.08
        elif ano and ano_candidato and abs(ano_candidato - ano) > 1:
            score -= 0.12

        if score > melhor_score:
            melhor_score = score
            melhor_doi = limpar_doi(item.get("DOI"))

    if melhor_score >= 0.90:
        return melhor_doi
    if melhor_score >= 0.84 and ano is not None:
        return melhor_doi
    return ""


def buscar_openalex_por_doi(doi: str, api_key: str | None) -> dict[str, Any]:
    external_id = quote(f"https://doi.org/{doi}", safe="")
    params = {"api_key": api_key} if api_key else None
    return request_json(f"https://api.openalex.org/works/{external_id}", params=params)


def urls_oa_candidatas(openalex_work: dict[str, Any]) -> list[str]:
    urls: list[str] = []

    best = openalex_work.get("best_oa_location") or {}
    primary = openalex_work.get("primary_location") or {}
    open_access = openalex_work.get("open_access") or {}

    for candidate in (
        best.get("pdf_url"),
        open_access.get("oa_url"),
        best.get("landing_page_url"),
        primary.get("pdf_url"),
        primary.get("landing_page_url"),
    ):
        if candidate and candidate not in urls:
            urls.append(candidate)

    return urls


def headers_pdf(url: str) -> dict[str, str]:
    parsed = urlparse(url)
    origem = f"{parsed.scheme}://{parsed.netloc}" if parsed.scheme and parsed.netloc else url
    return {
        **BROWSER_HEADERS,
        "Accept": "application/pdf,application/octet-stream;q=0.9,*/*;q=0.8",
        "Referer": origem,
    }


def headers_html(url: str) -> dict[str, str]:
    parsed = urlparse(url)
    origem = f"{parsed.scheme}://{parsed.netloc}" if parsed.scheme and parsed.netloc else url
    return {
        **BROWSER_HEADERS,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Referer": origem,
    }


def ler_resposta(url: str, headers: dict[str, str], timeout: int = 60) -> tuple[str, str, bytes]:
    request = Request(url, headers=headers)
    with urlopen(request, timeout=timeout) as response:
        conteudo = response.read()
        return response.geturl(), (response.headers.get("Content-Type") or "").lower(), conteudo


def extrair_urls_pdf_do_html(html: str, base_url: str) -> list[str]:
    candidatas: list[str] = []

    padroes = [
        r'<meta[^>]+name=["\']citation_pdf_url["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+name=["\']citation_pdf_url["\']',
        r'<meta[^>]+property=["\']og:url["\'][^>]+content=["\']([^"\']+\.pdf[^"\']*)["\']',
        r'href=["\']([^"\']+\.pdf(?:\?[^"\']*)?)["\']',
        r'href=["\']([^"\']+/pdf(?:\?[^"\']*)?)["\']',
        r'href=["\']([^"\']+/download(?:\?[^"\']*)?)["\']',
        r'"pdf_url"\s*:\s*"([^"]+)"',
    ]

    for padrao in padroes:
        for match in re.findall(padrao, html, flags=re.IGNORECASE):
            url = urljoin(base_url, match.replace("\\/", "/"))
            if url not in candidatas:
                candidatas.append(url)

    return candidatas


def resolver_urls_partindo_de_landing_page(url: str) -> list[str]:
    try:
        final_url, content_type, conteudo = ler_resposta(url, headers_html(url), timeout=20)
    except (HTTPError, URLError, TimeoutError, ValueError):
        return []

    if "pdf" in content_type or conteudo[:5] == b"%PDF-":
        return [final_url]

    html = conteudo.decode("utf-8", errors="ignore")
    return extrair_urls_pdf_do_html(html, final_url)


def expandir_urls_candidatas(urls: list[str], doi: str) -> list[str]:
    fila: list[str] = []
    vistas: set[str] = set()

    for url in [*urls, f"https://doi.org/{doi}"]:
        if url and url not in vistas:
            vistas.add(url)
            fila.append(url)

    expandidas: list[str] = []
    for url in fila:
        if url not in expandidas:
            expandidas.append(url)
        if ".pdf" in url.lower() or "/pdf" in url.lower():
            continue
        for pdf_url in resolver_urls_partindo_de_landing_page(url):
            if pdf_url not in expandidas:
                expandidas.append(pdf_url)

    return expandidas


def slug_titulo(texto: str, limite: int = 110) -> str:
    slug = normalizar(texto).replace(" ", "-")
    slug = re.sub(r"-+", "-", slug).strip("-")
    if not slug:
        return "sem-titulo"
    return slug[:limite].rstrip("-")


def nome_arquivo_pdf(titulo: str, autores: str, ano: int | None) -> str:
    prefixo = str(ano) if ano else "sano"
    return f"{prefixo}_{autor_principal(autores)}_{slug_titulo(titulo)}.pdf"


def baixar_pdf(url: str, destino: Path) -> bool:
    request = Request(url, headers=headers_pdf(url))
    with urlopen(request, timeout=25) as response:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            shutil.copyfileobj(response, tmp)
            tmp_path = Path(tmp.name)

    try:
        header = tmp_path.read_bytes()[:5]
        content_type = (response.headers.get("Content-Type") or "").lower()
        if "pdf" in content_type or header == b"%PDF-":
            destino.parent.mkdir(parents=True, exist_ok=True)
            shutil.move(str(tmp_path), destino)
            return True
    finally:
        if tmp_path.exists():
            tmp_path.unlink(missing_ok=True)

    return False


def mapear_colunas(worksheet) -> dict[str, int]:
    return {
        worksheet.cell(1, coluna).value: coluna
        for coluna in range(1, worksheet.max_column + 1)
        if worksheet.cell(1, coluna).value
    }


def garantir_colunas(worksheet) -> dict[str, int]:
    colunas = mapear_colunas(worksheet)
    for nome in NOVAS_COLUNAS:
        if nome not in colunas:
            coluna = worksheet.max_column + 1
            worksheet.cell(1, coluna).value = nome
            colunas[nome] = coluna
    return colunas


def obter_linhas(worksheet, aba: str, colunas: dict[str, int]) -> list[LinhaArtigo]:
    linhas: list[LinhaArtigo] = []
    for indice_linha in range(2, worksheet.max_row + 1):
        titulo = worksheet.cell(indice_linha, colunas[COLUNA_TITULO]).value
        if not titulo:
            continue
        ano_raw = worksheet.cell(indice_linha, colunas[COLUNA_ANO]).value
        doi_col = colunas.get("DOI")
        doi = worksheet.cell(indice_linha, doi_col).value if doi_col else ""
        try:
            ano = int(ano_raw) if ano_raw not in (None, "") else None
        except (TypeError, ValueError):
            ano = None
        linhas.append(
            LinhaArtigo(
                aba=aba,
                indice_linha=indice_linha,
                titulo=str(titulo).strip(),
                autores=str(worksheet.cell(indice_linha, colunas[COLUNA_AUTORES]).value or "").strip(),
                ano=ano,
                doi=limpar_doi(doi),
            )
        )
    return linhas


def atualizar_status_download(
    worksheet,
    colunas: dict[str, int],
    linha: LinhaArtigo,
    doi: str,
    status: str,
    url_pdf: str,
    arquivo_local: str,
) -> None:
    worksheet.cell(linha.indice_linha, colunas["DOI"]).value = doi or ""
    worksheet.cell(linha.indice_linha, colunas["Status do download OA"]).value = status
    worksheet.cell(linha.indice_linha, colunas["URL do PDF OA"]).value = url_pdf or ""
    worksheet.cell(linha.indice_linha, colunas["Arquivo local OA"]).value = arquivo_local or ""
    if status in {STATUS_BAIXADO, STATUS_EXISTENTE}:
        worksheet.cell(linha.indice_linha, colunas[COLUNA_BAIXADO]).value = "Sim"


def processar_linha(
    worksheet,
    colunas: dict[str, int],
    linha: LinhaArtigo,
    dois_locais: dict[str, str],
    pdfs_locais: list[PdfLocal],
    dry_run: bool,
    forcar: bool,
    mailto: str | None,
    api_key: str | None,
) -> str:
    doi = linha.doi or dois_locais.get(normalizar(linha.titulo), "")
    arquivo_local_existente = localizar_pdf_local_existente(linha.titulo, pdfs_locais)
    if arquivo_local_existente and not forcar:
        atualizar_status_download(
            worksheet,
            colunas,
            linha,
            doi,
            STATUS_EXISTENTE,
            worksheet.cell(linha.indice_linha, colunas["URL do PDF OA"]).value or "",
            str(arquivo_local_existente.relative_to(ROOT)),
        )
        return STATUS_EXISTENTE

    if not doi:
        doi = buscar_doi_crossref(linha.titulo, linha.ano, mailto)

    if not doi:
        atualizar_status_download(worksheet, colunas, linha, "", STATUS_SEM_DOI, "", "")
        return STATUS_SEM_DOI

    destino = PASTA_PDFS / nome_arquivo_pdf(linha.titulo, linha.autores, linha.ano)

    if destino.exists() and not forcar:
        atualizar_status_download(
            worksheet,
            colunas,
            linha,
            doi,
            STATUS_EXISTENTE,
            worksheet.cell(linha.indice_linha, colunas["URL do PDF OA"]).value or "",
            str(destino.relative_to(ROOT)),
        )
        return STATUS_EXISTENTE

    try:
        openalex_work = buscar_openalex_por_doi(doi, api_key)
    except (HTTPError, URLError, TimeoutError, json.JSONDecodeError):
        atualizar_status_download(worksheet, colunas, linha, doi, STATUS_ERRO, "", "")
        return STATUS_ERRO

    urls = expandir_urls_candidatas(urls_oa_candidatas(openalex_work), doi)
    if not urls:
        atualizar_status_download(worksheet, colunas, linha, doi, STATUS_SEM_PDF_OA, f"https://doi.org/{doi}", "")
        return STATUS_SEM_PDF_OA

    if dry_run:
        return "prévia"

    ultimo_url = ""
    for url in urls:
        ultimo_url = url
        try:
            if baixar_pdf(url, destino):
                atualizar_status_download(
                    worksheet,
                    colunas,
                    linha,
                    doi,
                    STATUS_BAIXADO,
                    url,
                    str(destino.relative_to(ROOT)),
                )
                return STATUS_BAIXADO
        except (HTTPError, URLError, TimeoutError, ValueError):
            continue

    atualizar_status_download(worksheet, colunas, linha, doi, STATUS_SEM_PDF_OA, ultimo_url, "")
    return STATUS_SEM_PDF_OA


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Baixa artigos open access e atualiza o catálogo em XLSX."
    )
    parser.add_argument(
        "--aba",
        choices=("artigos", "papers_useful", "todas"),
        default=ABA_PADRAO,
        help="Aba da planilha a processar.",
    )
    parser.add_argument(
        "--limite",
        type=int,
        default=None,
        help="Número máximo de linhas com título para processar.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Não baixa PDFs; apenas preenche DOI e prévia do que seria tentado.",
    )
    parser.add_argument(
        "--forcar",
        action="store_true",
        help="Rebaixa mesmo se o arquivo de destino já existir.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    workbook = load_workbook(PLANILHA)
    dois_locais = carregar_dois_locais()
    pdfs_locais = inventariar_pdfs_locais()
    mailto = os.getenv("CROSSREF_MAILTO")
    api_key = os.getenv("OPENALEX_API_KEY")

    abas = ("artigos", "papers_useful") if args.aba == "todas" else (args.aba,)
    total_processadas = 0
    resumo: dict[str, int] = {}

    for aba in abas:
        worksheet = workbook[aba]
        colunas = garantir_colunas(worksheet)
        linhas = obter_linhas(worksheet, aba, colunas)
        if not args.forcar:
            col_status = colunas.get("Status do download OA")
            if col_status:
                linhas = [
                    linha
                    for linha in linhas
                    if worksheet.cell(linha.indice_linha, col_status).value
                    not in {STATUS_BAIXADO, STATUS_EXISTENTE}
                ]
        if args.limite is not None:
            linhas = linhas[: args.limite]

        for linha in linhas:
            status = processar_linha(
                worksheet,
                colunas,
                linha,
                dois_locais,
                pdfs_locais,
                dry_run=args.dry_run,
                forcar=args.forcar,
                mailto=mailto,
                api_key=api_key,
            )
            resumo[status] = resumo.get(status, 0) + 1
            total_processadas += 1
            print(f"[{aba}] linha {linha.indice_linha}: {status} :: {linha.titulo}", flush=True)

    workbook.save(PLANILHA)

    print("\nResumo:")
    for status, quantidade in sorted(resumo.items()):
        print(f"- {status}: {quantidade}")
    print(f"- linhas processadas: {total_processadas}")

    if not api_key:
        print(
            "\nAviso: OPENALEX_API_KEY não definido. Para lotes maiores, configure uma chave "
            "gratuita em https://openalex.org/settings/api .",
            file=sys.stderr,
        )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
