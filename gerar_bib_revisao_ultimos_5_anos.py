#!/usr/bin/env python3
"""
Gera um arquivo .bib consolidado para uso no bibliometrix a partir de:

- aba `papers_useful` da planilha `bibliography/papers/catalogo_artigos.xlsx`
- entradas recentes de `bibliography/reference.bib`

Regra padrao de recorte:
- 5 anos correntes, incluindo o ano atual do ambiente
- exemplo em 2026: 2022-2026

Quando houver sobreposicao entre planilha e `reference.bib`, a entrada mais rica
do `reference.bib` e mantida, e a entrada minima gerada da planilha e descartada.
"""

from __future__ import annotations

import argparse
import re
import unicodedata
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


RAIZ = Path(__file__).resolve().parents[1]
PLANILHA = RAIZ / "bibliography" / "papers" / "catalogo_artigos.xlsx"
ARQUIVO_REFERENCIAS = RAIZ / "bibliography" / "reference.bib"
ABA_UTIL = "papers_useful"


@dataclass
class EntradaBib:
    chave: str
    ano: int
    titulo: str
    doi_norm: str
    titulo_norm: str
    origem: str
    bruto: str


def ascii_slug(texto: str) -> str:
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = texto.lower()
    texto = re.sub(r"[^a-z0-9]+", "-", texto)
    return texto.strip("-")


def normalizar_doi(valor: str | None) -> str:
    if not valor:
        return ""
    valor = str(valor).strip().lower()
    valor = re.sub(r"^https?://(dx\.)?doi\.org/", "", valor)
    return valor.strip()


def normalizar_titulo(valor: str | None) -> str:
    if not valor:
        return ""
    valor = unicodedata.normalize("NFKD", str(valor))
    valor = "".join(ch for ch in valor if not unicodedata.combining(ch))
    valor = valor.lower()
    valor = re.sub(r"[^a-z0-9]+", " ", valor)
    return re.sub(r"\s+", " ", valor).strip()


def valor_bib(valor: str | None) -> str:
    if valor is None:
        return ""
    texto = str(valor).replace("\n", " ").strip()
    texto = re.sub(r"\s+", " ", texto)
    texto = texto.replace("{", "\\{").replace("}", "\\}")
    return texto


def normalizar_layout_bib(bruto: str) -> str:
    return "\n".join(linha.lstrip() for linha in bruto.strip().splitlines())


def inserir_campo_se_ausente(bruto: str, campo: str, valor: str) -> str:
    if re.search(rf"(?im)^{re.escape(campo)}\s*=", bruto):
        return bruto
    linhas = bruto.strip().splitlines()
    if len(linhas) < 2:
        return bruto
    linhas.insert(-1, f"{campo} = {{{valor_bib(valor)}}},")
    return "\n".join(linhas)


def sobrenome_primeiro_autor(autores: str) -> str:
    primeiro = (autores or "sem_autor").split(";")[0].strip()
    if "," in primeiro:
        return primeiro.split(",", 1)[0].strip() or "sem_autor"
    partes = primeiro.split()
    return partes[-1] if partes else "sem_autor"


def criar_chave(autores: str, ano: int, titulo: str, chaves_existentes: set[str]) -> str:
    base_autor = ascii_slug(sobrenome_primeiro_autor(autores)) or "semautor"
    palavras = [p for p in ascii_slug(titulo).split("-") if p]
    base_titulo = "".join(palavras[:3]) or "artigo"
    base = f"{base_autor}{ano}{base_titulo}"
    chave = base
    contador = 2
    while chave in chaves_existentes:
        chave = f"{base}{contador}"
        contador += 1
    return chave


def encontrar_virgula_nivel_zero(texto: str) -> int:
    nivel = 0
    for i, ch in enumerate(texto):
        if ch == "{":
            nivel += 1
        elif ch == "}":
            nivel -= 1
        elif ch == "," and nivel == 0:
            return i
    raise ValueError("Entrada BibTeX invalida: chave nao encontrada.")


def extrair_valor_campo(texto: str, campo: str) -> str | None:
    padrao = re.compile(rf"\b{re.escape(campo)}\s*=\s*", flags=re.IGNORECASE)
    match = padrao.search(texto)
    if not match:
        return None

    i = match.end()
    while i < len(texto) and texto[i].isspace():
        i += 1
    if i >= len(texto):
        return None

    if texto[i] == "{":
        nivel = 0
        inicio = i + 1
        i += 1
        while i < len(texto):
            if texto[i] == "{":
                nivel += 1
            elif texto[i] == "}":
                if nivel == 0:
                    return texto[inicio:i].strip()
                nivel -= 1
            i += 1
        return None

    if texto[i] == '"':
        inicio = i + 1
        i += 1
        while i < len(texto):
            if texto[i] == '"' and texto[i - 1] != "\\":
                return texto[inicio:i].strip()
            i += 1
        return None

    inicio = i
    while i < len(texto) and texto[i] not in ",\n":
        i += 1
    return texto[inicio:i].strip()


def entrada_reference_parece_valida(campos: str) -> bool:
    titulo = extrair_valor_campo(campos, "title") or ""
    journal = extrair_valor_campo(campos, "journal") or ""
    doi = extrair_valor_campo(campos, "doi") or ""

    titulo_norm = normalizar_titulo(titulo)
    journal_norm = normalizar_titulo(journal)
    doi_norm = normalizar_doi(doi)

    if doi_norm and "xxxxx" in doi_norm:
        return False
    if len(titulo_norm.split()) < 3:
        return False
    if journal_norm and titulo_norm == journal_norm:
        return False
    return True


def iterar_entradas_bib(texto: str) -> Iterable[str]:
    i = 0
    tamanho = len(texto)
    while i < tamanho:
        inicio = texto.find("@", i)
        if inicio == -1:
            break
        brace = texto.find("{", inicio)
        if brace == -1:
            break
        nivel = 0
        fim = None
        for j in range(brace, tamanho):
            if texto[j] == "{":
                nivel += 1
            elif texto[j] == "}":
                nivel -= 1
                if nivel == 0:
                    fim = j + 1
                    break
        if fim is None:
            break
        yield texto[inicio:fim].strip()
        i = fim


def carregar_reference_bib(caminho: Path, ano_inicial: int, ano_final: int) -> list[EntradaBib]:
    texto = caminho.read_text(encoding="utf-8")
    entradas: list[EntradaBib] = []
    vistos: set[tuple[str, str]] = set()

    for bruto in iterar_entradas_bib(texto):
        tipo_sep = bruto.find("{")
        miolo = bruto[tipo_sep + 1 : -1].strip()
        try:
            idx_virgula = encontrar_virgula_nivel_zero(miolo)
        except ValueError:
            continue

        chave = miolo[:idx_virgula].strip()
        campos = miolo[idx_virgula + 1 :]
        if not entrada_reference_parece_valida(campos):
            continue

        ano_str = extrair_valor_campo(campos, "year")
        titulo = extrair_valor_campo(campos, "title") or ""
        doi = extrair_valor_campo(campos, "doi") or ""

        if not ano_str or not ano_str.isdigit():
            continue
        ano = int(ano_str)
        if not (ano_inicial <= ano <= ano_final):
            continue

        doi_norm = normalizar_doi(doi)
        titulo_norm = normalizar_titulo(titulo)
        marcador = (doi_norm, titulo_norm)
        if marcador in vistos:
            continue
        vistos.add(marcador)

        bruto_norm = normalizar_layout_bib(bruto)
        bruto_norm = inserir_campo_se_ausente(bruto_norm, "affiliation", "NA")
        bruto_norm = inserir_campo_se_ausente(bruto_norm, "references", "none")

        entradas.append(
            EntradaBib(
                chave=chave,
                ano=ano,
                titulo=titulo,
                doi_norm=doi_norm,
                titulo_norm=titulo_norm,
                origem="reference.bib",
                bruto=bruto_norm,
            )
        )

    return entradas


def carregar_planilha(caminho: Path, aba: str, ano_inicial: int, ano_final: int) -> list[dict[str, str]]:
    wb = load_workbook(caminho, read_only=True, data_only=True)
    ws = wb[aba]
    cabecalho = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    indice = {nome: i for i, nome in enumerate(cabecalho)}

    linhas: list[dict[str, str]] = []
    for linha in ws.iter_rows(min_row=2, values_only=True):
        ano_raw = linha[indice["Ano de publicação"]]
        if not isinstance(ano_raw, (int, float)):
            continue
        ano = int(ano_raw)
        if not (ano_inicial <= ano <= ano_final):
            continue

        linhas.append(
            {
                "titulo": str(linha[indice["Título do artigo"]] or "").strip(),
                "autores": str(linha[indice["Autores"]] or "").strip(),
                "ano": str(ano),
                "palavras_chave": str(linha[indice["Palavras-chave"]] or "").strip(),
                "revista": str(linha[indice["Nome da revista"]] or "").strip(),
                "doi": str(linha[indice["DOI"]] or "").strip(),
                "url": str(linha[indice["URL do PDF OA"]] or "").strip(),
                "arquivo_local": str(linha[indice["Arquivo local OA"]] or "").strip(),
            }
        )
    return linhas


def gerar_entrada_minima(
    linha: dict[str, str],
    chaves_existentes: set[str],
    data_geracao: str,
) -> EntradaBib:
    ano = int(linha["ano"])
    chave = criar_chave(linha["autores"], ano, linha["titulo"], chaves_existentes)
    campos = [
        ("author", linha["autores"]),
        ("title", linha["titulo"]),
        ("journal", linha["revista"]),
        ("year", linha["ano"]),
        ("doi", linha["doi"]),
        ("url", linha["url"]),
        ("affiliation", "NA"),
        ("references", "none"),
        ("author_keywords", linha["palavras_chave"]),
        ("keywords", linha["palavras_chave"]),
        (
            "note",
            f"Importado automaticamente da aba {ABA_UTIL} em {data_geracao}"
            + (f"; arquivo local: {linha['arquivo_local']}" if linha["arquivo_local"] else ""),
        ),
    ]

    linhas = [f"@article{{{chave},"]
    for nome, valor in campos:
        if valor:
            linhas.append(f"{nome} = {{{valor_bib(valor)}}},")
    linhas.append("}")

    return EntradaBib(
        chave=chave,
        ano=ano,
        titulo=linha["titulo"],
        doi_norm=normalizar_doi(linha["doi"]),
        titulo_norm=normalizar_titulo(linha["titulo"]),
        origem="planilha",
        bruto=normalizar_layout_bib("\n".join(linhas)),
    )


def montar_saida(
    entradas_reference: list[EntradaBib],
    linhas_planilha: list[dict[str, str]],
    ano_inicial: int,
    ano_final: int,
) -> tuple[str, dict[str, int]]:
    data_geracao = date.today().isoformat()
    chaves_existentes = {entrada.chave for entrada in entradas_reference}
    marcadores_existentes = {
        (entrada.doi_norm, entrada.titulo_norm) for entrada in entradas_reference
    }

    entradas_finais = list(entradas_reference)
    descartadas_por_sobreposicao = 0

    for linha in linhas_planilha:
        marcador = (
            normalizar_doi(linha["doi"]),
            normalizar_titulo(linha["titulo"]),
        )
        if marcador in marcadores_existentes:
            descartadas_por_sobreposicao += 1
            continue

        entrada = gerar_entrada_minima(linha, chaves_existentes, data_geracao)
        chaves_existentes.add(entrada.chave)
        marcadores_existentes.add((entrada.doi_norm, entrada.titulo_norm))
        entradas_finais.append(entrada)

    entradas_finais.sort(key=lambda item: (-item.ano, normalizar_titulo(item.titulo), item.chave))

    cabecalho = [
        "% ============================================================",
        f"% Consolidado para bibliometrix: {ano_inicial}-{ano_final}",
        f"% Fontes: {PLANILHA.relative_to(RAIZ)} (aba {ABA_UTIL}) + {ARQUIVO_REFERENCIAS.relative_to(RAIZ)}",
        f"% Gerado automaticamente em {data_geracao}",
        "% Regra de deduplicacao: DOI normalizado; fallback por titulo normalizado",
        "% ============================================================",
        "",
    ]

    conteudos = cabecalho + [entrada.bruto + "\n" for entrada in entradas_finais]
    estatisticas = {
        "planilha_filtrada": len(linhas_planilha),
        "reference_filtrada": len(entradas_reference),
        "sobreposicoes_descartadas": descartadas_por_sobreposicao,
        "total_final": len(entradas_finais),
    }
    return "\n".join(conteudos).rstrip() + "\n", estatisticas


def main() -> None:
    ano_final_padrao = date.today().year
    ano_inicial_padrao = ano_final_padrao - 4

    parser = argparse.ArgumentParser(
        description="Gera um .bib consolidado para revisao bibliografica recente."
    )
    parser.add_argument("--ano-inicial", type=int, default=ano_inicial_padrao)
    parser.add_argument("--ano-final", type=int, default=ano_final_padrao)
    parser.add_argument("--saida", type=Path, default=None)
    args = parser.parse_args()

    if args.ano_inicial > args.ano_final:
        raise SystemExit("--ano-inicial nao pode ser maior que --ano-final.")

    if args.saida is None:
        args.saida = (
            RAIZ
            / "bibliography"
            / f"revisao_bibliografica_{args.ano_inicial}_{args.ano_final}.bib"
        )

    entradas_reference = carregar_reference_bib(
        ARQUIVO_REFERENCIAS, args.ano_inicial, args.ano_final
    )
    linhas_planilha = carregar_planilha(
        PLANILHA, ABA_UTIL, args.ano_inicial, args.ano_final
    )
    conteudo, estatisticas = montar_saida(
        entradas_reference,
        linhas_planilha,
        args.ano_inicial,
        args.ano_final,
    )

    args.saida.write_text(conteudo, encoding="utf-8")

    print(f"Arquivo gerado: {args.saida.relative_to(RAIZ)}")
    print(f"Recorte: {args.ano_inicial}-{args.ano_final}")
    print(f"Entradas da planilha no recorte: {estatisticas['planilha_filtrada']}")
    print(f"Entradas do reference.bib no recorte: {estatisticas['reference_filtrada']}")
    print(f"Sobreposicoes descartadas: {estatisticas['sobreposicoes_descartadas']}")
    print(f"Total final no .bib: {estatisticas['total_final']}")


if __name__ == "__main__":
    main()
